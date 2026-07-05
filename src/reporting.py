import json
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from src.config import (
    LATEST_EXECUTION_EXCEL,
    MARGEN_BUENO,
    MARGEN_MUY_BUENO,
    OUTPUT_PARQUET,
    REPORT_CSV,
    REPORT_EXCEL,
    REPORT_JSON,
    REPORTS_DIR,
)
from src.transformer import build_item_key


DETAIL_COLUMNS = [
    "item_key",
    "expansion",
    "rareza",
    "nombre_carta",
    "numero_carta",
    "precio_anterior",
    "precio_actual",
    "diferencia_precio",
    "porcentaje_variacion",
    "estado_variacion",
    "precio_referencia",
    "margen_estimado",
    "porcentaje_margen",
    "clasificacion_oportunidad",
    "estado_scraping",
    "mensaje_error",
    "timestamp_ejecucion",
    "url_carta",
]


def _with_item_key(df: pd.DataFrame) -> pd.DataFrame:
    output = df.copy()
    if "item_key" not in output.columns or output["item_key"].fillna("").astype(str).str.strip().eq("").any():
        output["item_key"] = build_item_key(output)
    return output


def load_historical_data(path: Path = OUTPUT_PARQUET) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"No existe el Parquet historico: {path}")
    df = _with_item_key(pd.read_parquet(path))
    if df.empty:
        raise ValueError("El Parquet historico esta vacio.")
    return df


def _timestamp_values(df: pd.DataFrame) -> pd.Series:
    if "timestamp_ejecucion" not in df.columns:
        raise ValueError("Falta la columna timestamp_ejecucion en el historico.")
    timestamps = pd.to_datetime(df["timestamp_ejecucion"], errors="coerce")
    if timestamps.dropna().empty:
        raise ValueError("No hay timestamps validos para comparar ejecuciones.")
    return timestamps


def get_latest_execution(df: pd.DataFrame) -> pd.DataFrame:
    timestamps = _timestamp_values(df)
    return df[timestamps.eq(timestamps.max())].copy()


def get_previous_execution(df: pd.DataFrame) -> pd.DataFrame:
    timestamps = _timestamp_values(df)
    latest = timestamps.max()
    previous_values = timestamps[timestamps.lt(latest)]
    if previous_values.dropna().empty:
        return pd.DataFrame(columns=df.columns)
    return df[timestamps.eq(previous_values.max())].copy()


def _last_by_item(df: pd.DataFrame, price_col: str) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["item_key", price_col])
    return (
        _with_item_key(df)
        .drop_duplicates(subset=["item_key"], keep="last")[["item_key", "market_price_usd"]]
        .rename(columns={"market_price_usd": price_col})
    )


def classify_price_change(row: pd.Series) -> str:
    if str(row.get("estado_scraping", "")).upper() == "ERROR":
        return "ERROR"
    if pd.isna(row.get("precio_actual")):
        return "SIN_PRECIO"
    if pd.isna(row.get("precio_anterior")):
        return "NUEVO"
    if row["precio_actual"] < row["precio_anterior"]:
        return "BAJO"
    if row["precio_actual"] > row["precio_anterior"]:
        return "SUBIO"
    return "IGUAL"


def _classify_opportunity(row: pd.Series) -> str:
    if pd.isna(row.get("precio_referencia")) or pd.isna(row.get("precio_actual")) or row["precio_referencia"] <= 0:
        return "SIN_REFERENCIA"
    if row["porcentaje_margen"] >= MARGEN_MUY_BUENO:
        return "MUY_BUENA_OPORTUNIDAD"
    if row["porcentaje_margen"] >= MARGEN_BUENO:
        return "BUENA_OPORTUNIDAD"
    if row["porcentaje_margen"] > 0:
        return "NORMAL"
    return "NO_CONVIENE"


def compare_prices(latest_df: pd.DataFrame, previous_df: pd.DataFrame) -> pd.DataFrame:
    latest = _with_item_key(latest_df).drop_duplicates(subset=["item_key"], keep="last").copy()
    previous = _last_by_item(previous_df, "precio_anterior")

    latest["precio_actual"] = pd.to_numeric(latest.get("market_price_usd"), errors="coerce")
    latest["precio_referencia"] = pd.to_numeric(latest.get("precio_referencia"), errors="coerce")
    report = latest.merge(previous, on="item_key", how="left")
    report["precio_anterior"] = pd.to_numeric(report["precio_anterior"], errors="coerce")
    report["diferencia_precio"] = report["precio_actual"] - report["precio_anterior"]
    report["porcentaje_variacion"] = report["diferencia_precio"] / report["precio_anterior"]
    report.loc[report["precio_anterior"].eq(0), "porcentaje_variacion"] = pd.NA
    report["estado_variacion"] = report.apply(classify_price_change, axis=1)
    report["margen_estimado"] = report["precio_referencia"] - report["precio_actual"]
    report["porcentaje_margen"] = report["margen_estimado"] / report["precio_referencia"]
    report.loc[report["precio_referencia"].eq(0), "porcentaje_margen"] = pd.NA
    report["clasificacion_oportunidad"] = report.apply(_classify_opportunity, axis=1)

    for col in DETAIL_COLUMNS:
        if col not in report.columns:
            report[col] = None
    return report[DETAIL_COLUMNS].sort_values(["estado_variacion", "nombre_carta"], na_position="last")


def build_summary(detail_df: pd.DataFrame) -> dict:
    counts = detail_df["estado_variacion"].value_counts()
    opportunities = detail_df["clasificacion_oportunidad"].isin(
        ["MUY_BUENA_OPORTUNIDAD", "BUENA_OPORTUNIDAD"]
    )
    latest_timestamp = ""
    if "timestamp_ejecucion" in detail_df.columns and not detail_df.empty:
        latest_timestamp = str(detail_df["timestamp_ejecucion"].dropna().max())

    return {
        "fecha_ejecucion": latest_timestamp,
        "total_items": int(len(detail_df)),
        "cantidad_bajaron": int(counts.get("BAJO", 0)),
        "cantidad_subieron": int(counts.get("SUBIO", 0)),
        "cantidad_igual": int(counts.get("IGUAL", 0)),
        "cantidad_nuevos": int(counts.get("NUEVO", 0)),
        "cantidad_error": int(counts.get("ERROR", 0)),
        "cantidad_sin_precio": int(counts.get("SIN_PRECIO", 0)),
        "mejores_oportunidades": int(opportunities.sum()),
    }


def _format_sheet(writer, sheet_name: str) -> None:
    sheet = writer.book[sheet_name]
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_letter].width = min(max(max_len + 2, 12), 45)


def export_latest_execution_excel(df_latest: pd.DataFrame, path: Path = LATEST_EXECUTION_EXCEL) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    output = _with_item_key(df_latest)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        output.to_excel(writer, index=False, sheet_name="Ultima ejecucion")
        _format_sheet(writer, "Ultima ejecucion")

    pd.read_excel(path, sheet_name="Ultima ejecucion")
    return path


def generate_report_files(
    detail_df: pd.DataFrame,
    summary: dict,
    excel_path: Path = REPORT_EXCEL,
    csv_path: Path = REPORT_CSV,
    json_path: Path = REPORT_JSON,
) -> tuple[Path, Path, Path]:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    detail_df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    json_path.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")

    bajaron = detail_df[detail_df["estado_variacion"].eq("BAJO")].sort_values("diferencia_precio")
    subieron = detail_df[detail_df["estado_variacion"].eq("SUBIO")].sort_values("diferencia_precio", ascending=False)
    oportunidades = detail_df[
        detail_df["clasificacion_oportunidad"].isin(["MUY_BUENA_OPORTUNIDAD", "BUENA_OPORTUNIDAD"])
    ].sort_values("porcentaje_margen", ascending=False)

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        pd.DataFrame([summary]).to_excel(writer, index=False, sheet_name="Resumen")
        detail_df.to_excel(writer, index=False, sheet_name="Detalle")
        bajaron.to_excel(writer, index=False, sheet_name="Bajaron")
        subieron.to_excel(writer, index=False, sheet_name="Subieron")
        oportunidades.to_excel(writer, index=False, sheet_name="Oportunidades")
        for sheet_name in writer.book.sheetnames:
            _format_sheet(writer, sheet_name)

    pd.read_excel(excel_path, sheet_name="Resumen")
    return excel_path, csv_path, json_path


def generate_latest_report() -> tuple[pd.DataFrame, dict, tuple[Path, Path, Path]]:
    historical = load_historical_data()
    latest = get_latest_execution(historical)
    export_latest_execution_excel(latest)
    previous = get_previous_execution(historical)
    detail = compare_prices(latest, previous)
    summary = build_summary(detail)
    paths = generate_report_files(detail, summary)
    return detail, summary, paths
