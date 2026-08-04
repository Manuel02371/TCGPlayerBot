from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

DISPLAY_COLUMNS = {
    "expansion": "Expansión", "expansion_code": "Código de expansión",
    "card_name": "Nombre de carta", "card_number": "Número", "language": "Idioma",
    "seller_count": "Cantidad de vendedores", "price_from": "Precio desde",
    "market_price": "Precio de mercado", "catalog_url": "URL del catálogo de la expansión",
    "card_url": "URL de la carta", "execution_id": "execution_id", "extracted_at": "extracted_at",
    "execution_date": "Fecha de ejecución", "observation_key": "observation_key",
    "previous_market_price": "Precio anterior", "change_amount": "Variación", "change_percent": "Variación %",
    "status": "Estado", "previous_extracted_at": "Fecha del corte anterior", "prior_cuts_used": "Cortes previos usados",
    "price_cut_1": "Precio corte 1", "price_cut_2": "Precio corte 2", "price_cut_3": "Precio corte 3",
    "price_cut_4": "Precio corte 4", "price_cut_5": "Precio corte 5",
}


def write_excel(frame: pd.DataFrame, path: Path, sheet_name: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    output = frame.rename(columns=DISPLAY_COLUMNS).copy()
    output.to_excel(path, index=False, sheet_name=sheet_name)
    book = load_workbook(path)
    sheet = book[sheet_name]
    sheet.freeze_panes = "A2"
    sheet.row_dimensions[1].height = 22
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    if sheet.max_column and sheet.max_row:
        table = Table(displayName="TablaPrecios", ref=sheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        sheet.add_table(table)
    for index, cell in enumerate(sheet[1], start=1):
        values = [str(cell.value or "")] + [str(row[0] or "") for row in sheet.iter_rows(min_row=2, min_col=index, max_col=index)]
        sheet.column_dimensions[cell.column_letter].width = min(max(map(len, values)) + 2, 48)
        if cell.value in {"Precio desde", "Precio de mercado", "Precio anterior", "Variación", "Precio corte 1", "Precio corte 2", "Precio corte 3", "Precio corte 4", "Precio corte 5"}:
            for price_cell in sheet.iter_cols(min_col=index, max_col=index, min_row=2):
                for value in price_cell:
                    value.number_format = '#,##0.00'
        if cell.value == "Variación %":
            for percent_cell in sheet.iter_cols(min_col=index, max_col=index, min_row=2):
                for value in percent_cell:
                    value.number_format = '0.00%'
    book.save(path)
