from __future__ import annotations

import re
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parent
STOCK_PATH = ROOT / "data" / "input" / "Input stock.xlsx"
CURRENT_PATH = ROOT / "data" / "output" / "precios_actuales.xlsx"
HISTORY_PATH = ROOT / "data" / "output" / "historico_precios.parquet"
OUTPUT_PATH = ROOT / "data" / "output" / "precios_a_modificar.xlsx"


def normalized(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value)).encode("ascii", "ignore").decode().lower()
    return re.sub(r"[^a-z0-9]+", "", text)


def number_from_name(name: object) -> str:
    found = re.search(r"(\d{1,4}/\d{1,4})\s*$", str(name))
    return found.group(1) if found else ""


def clean_card_name(name: object) -> str:
    return re.sub(r"\s+\d{1,4}/\d{1,4}\s*$", "", str(name)).strip()


def card_key(expansion: object, name: object, number: object) -> str:
    return "|".join((normalized(expansion), normalized(clean_card_name(name)), normalized(number)))


def price_recommendation(current_price: float, market_minimum: float, stock: int) -> dict[str, float | str]:
    floor = 0.15 if stock > 10 and market_minimum < 0.30 else 0.30
    suggested_price = max(floor, market_minimum)
    if current_price > market_minimum and suggested_price < current_price:
        decision = "Bajar"
    elif current_price < market_minimum and current_price < suggested_price:
        decision = "Subir"
    else:
        decision = "Mantener"
    return {"floor": floor, "suggested_price": suggested_price, "decision": decision}


def history_trends() -> dict[str, str]:
    if not HISTORY_PATH.exists():
        return {}
    history = pd.read_parquet(HISTORY_PATH).dropna(subset=["market_price", "extracted_at"])
    history["key"] = history.apply(lambda row: card_key(row["expansion"], row["card_name"], row["card_number"]), axis=1)
    history["extracted_at"] = pd.to_datetime(history["extracted_at"], errors="coerce")
    trends: dict[str, str] = {}
    for key, group in history.dropna(subset=["extracted_at"]).groupby("key"):
        runs = group.sort_values("extracted_at").drop_duplicates("execution_id", keep="last")
        if len(runs) < 2:
            trends[key] = "Sin histórico"
            continue
        latest, previous = runs.iloc[-1]["market_price"], runs.iloc[-2]["market_price"]
        trends[key] = "Subió" if latest > previous else "Bajó" if latest < previous else "Sin cambio"
    return trends


def read_inputs() -> tuple[pd.DataFrame, pd.DataFrame]:
    if not STOCK_PATH.exists():
        raise FileNotFoundError(f"Falta el stock: {STOCK_PATH}")
    if not CURRENT_PATH.exists():
        raise FileNotFoundError(f"Faltan los precios actuales: {CURRENT_PATH}")
    stock = pd.read_excel(STOCK_PATH, sheet_name="Catalogo")
    current = pd.read_excel(CURRENT_PATH, sheet_name="Precios actuales")
    if len(stock.columns) != 9 or len(current.columns) != 10:
        raise ValueError("Los Excel de stock o precios actuales no tienen las columnas esperadas.")
    stock.columns = ["card_name", "stock_number", "expansion_code", "expansion", "language", "condition", "variant", "stock", "current_price"]
    current.columns = ["expansion", "expansion_code", "card_name", "card_number", "language", "seller_count", "price_from", "market_price", "catalog_url", "card_url"]
    return stock, current


def build_reports(stock: pd.DataFrame, current: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    eligible = stock[
        stock["language"].map(normalized).eq("ingles")
        & stock["variant"].map(normalized).eq("normal")
    ].copy()
    eligible["card_number"] = eligible["card_name"].map(number_from_name)
    eligible["key"] = eligible.apply(lambda row: card_key(row["expansion"], row["card_name"], row["card_number"]), axis=1)
    current["key"] = current.apply(lambda row: card_key(row["expansion"], row["card_name"], row["card_number"]), axis=1)
    market = current.drop_duplicates("key", keep="first").set_index("key")
    trends = history_trends()
    matched, unmatched = [], []
    for item in eligible.to_dict("records"):
        base = {
            "Expansión": item["expansion"], "Carta": clean_card_name(item["card_name"]), "Número": item["card_number"],
            "Variante": item["variant"], "Stock": int(item["stock"]), "Condición": item["condition"], "Precio actual": float(item["current_price"]),
        }
        if item["key"] not in market.index:
            unmatched.append({**base, "Motivo": "No aparece en precios_actuales.xlsx; no se recomienda precio."})
            continue
        price = market.loc[item["key"]]
        recommendation = price_recommendation(base["Precio actual"], float(price["price_from"]), base["Stock"])
        matched.append({
            **base, "Vendedores": int(price["seller_count"]), "Precio mínimo mercado": float(price["price_from"]),
            "Precio mercado": float(price["market_price"]), "Tendencia mercado": trends.get(item["key"], "Sin histórico"),
            "Piso permitido": recommendation["floor"], "Precio sugerido": recommendation["suggested_price"],
            "Decisión": recommendation["decision"], "Prioridad": "Alta" if price["seller_count"] >= 10 else "Media" if price["seller_count"] >= 5 else "Baja",
            "Motivo": "", "URL carta": price["card_url"],
        })
    report = pd.DataFrame(matched)
    if not report.empty:
        report["Motivo"] = report.apply(
            lambda row: "Bajar: queda al mínimo del mercado" if row["Decisión"] == "Bajar" and row["Precio sugerido"] == row["Precio mínimo mercado"]
            else "Bajar: el piso impide superar el mínimo" if row["Decisión"] == "Bajar"
            else "Subir: mantiene el primer precio" if row["Decisión"] == "Subir"
            else "Mantener: ya está igual o por debajo del mínimo", axis=1,
        )
        report = report.sort_values(["Expansión", "Carta", "Número"])
    changes = report[report["Decisión"] != "Mantener"].copy() if not report.empty else report
    holds = report[report["Decisión"] == "Mantener"].copy() if not report.empty else report
    return changes, holds, pd.DataFrame(unmatched)


def style_sheet(writer: pd.ExcelWriter, sheet_name: str) -> None:
    sheet = writer.book[sheet_name]
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    sheet.auto_filter.ref = sheet.dimensions
    if sheet.max_row > 1:
        table = Table(displayName=f"Table{re.sub(r'[^A-Za-z0-9]', '', sheet_name)}", ref=sheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        sheet.add_table(table)
    for column in sheet.columns:
        letter = column[0].column_letter
        sheet.column_dimensions[letter].width = min(max(len(str(cell.value or "")) for cell in column) + 2, 48)
    for header in ("Precio actual", "Precio mínimo mercado", "Precio mercado", "Piso permitido", "Precio sugerido"):
        for cell in sheet[1]:
            if cell.value == header:
                for row in range(2, sheet.max_row + 1):
                    sheet.cell(row, cell.column).number_format = "#,##0.00"


def write_output(changes: pd.DataFrame, holds: pd.DataFrame, unmatched: pd.DataFrame) -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    lower = int((changes["Decisión"] == "Bajar").sum()) if "Decisión" in changes else 0
    raise_count = int((changes["Decisión"] == "Subir").sum()) if "Decisión" in changes else 0
    summary = pd.DataFrame({"Indicador": ["Precios a modificar", "Bajar", "Subir", "Mantener", "Sin match"], "Cantidad": [len(changes), lower, raise_count, len(holds), len(unmatched)]})
    methodology = pd.DataFrame({"Regla": ["Filtro", "Piso", "Igual al mínimo", "Precio por debajo de mercado"], "Aplicación": ["Inglés; Solo Normal", "S/ 0.30; S/ 0.15 como mínimo solo si Stock > 10 y mínimo mercado < S/ 0.30", "Mantener", "Aumentar al mínimo"]})
    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Resumen", index=False)
        changes.to_excel(writer, sheet_name="Precios a modificar", index=False)
        holds.to_excel(writer, sheet_name="Sin cambio", index=False)
        unmatched.to_excel(writer, sheet_name="Sin match", index=False)
        methodology.to_excel(writer, sheet_name="Metodología", index=False)
        for name in writer.book.sheetnames:
            style_sheet(writer, name)


def main() -> None:
    stock, current = read_inputs()
    changes, holds, unmatched = build_reports(stock, current)
    try:
        write_output(changes, holds, unmatched)
    except PermissionError as error:
        raise SystemExit(f"Cierre {OUTPUT_PATH.name} y vuelva a ejecutar el BAT.") from error
    print(f"Archivo generado: {OUTPUT_PATH}")
    print(f"Cambios: {len(changes)} | Sin cambio: {len(holds)} | Sin match: {len(unmatched)}")


if __name__ == "__main__":
    main()
